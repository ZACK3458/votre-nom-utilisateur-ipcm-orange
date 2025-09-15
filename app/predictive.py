"""
Module d'analyse prédictive de capacité selon spécifications IPCM.

Ce module implémente des algorithmes de prédiction pour anticiper les besoins
futurs en capacité réseau basés sur l'historique d'utilisation des interfaces.
Supporte différents modèles: régression linéaire, polynomiale, et moyennes mobiles.
"""

from typing import Iterable, List, Tuple, Dict, Any, Optional
from datetime import datetime, timedelta
import json
import math

from .inventory.store import load_interface_history
from .inventory.models import Equipment, Interface, NetworkDomain


def _linear_regression_coeffs(y: Iterable[float]) -> Tuple[float, float]:
    """Calcule les coefficients (pente m, ordonnée b) via moindres carrés.

    Utilise x = 0..n-1 pour les abscisses. Gère les cas dégénérés (n<2).
    """
    y_list = list(float(v) for v in y)
    n = len(y_list)
    if n < 2:
        # Pas assez de points pour une pente: retourne moyenne et pente nulle
        mean = y_list[0] if n == 1 else 0.0
        return 0.0, mean

    sum_x = (n - 1) * n / 2.0
    sum_x2 = (n - 1) * n * (2 * n - 1) / 6.0
    sum_y = sum(y_list)
    sum_xy = sum(i * v for i, v in enumerate(y_list))

    denom = n * sum_x2 - sum_x * sum_x
    if denom == 0:
        # Tous x identiques (ne devrait pas arriver ici) -> pente nulle
        m = 0.0
    else:
        m = (n * sum_xy - sum_x * sum_y) / denom
    b = (sum_y - m * sum_x) / n
    return m, b


def _polynomial_regression_coeffs(y: Iterable[float], degree: int = 2) -> List[float]:
    """Calcule les coefficients d'une régression polynomiale simple."""
    y_list = list(float(v) for v in y)
    n = len(y_list)
    
    if n < degree + 1:
        # Pas assez de points pour le degré demandé
        return [sum(y_list) / n if n > 0 else 0.0]
    
    # Pour une implémentation simple, on se limite à degré 2 (quadratique)
    if degree == 2 and n >= 3:
        # Régression quadratique: y = ax² + bx + c
        # Résolution par système d'équations normales simplifiée
        sum_x = sum(range(n))
        sum_x2 = sum(i**2 for i in range(n))
        sum_x3 = sum(i**3 for i in range(n))
        sum_x4 = sum(i**4 for i in range(n))
        sum_y = sum(y_list)
        sum_xy = sum(i * y_list[i] for i in range(n))
        sum_x2y = sum(i**2 * y_list[i] for i in range(n))
        
        # Matrice 3x3 et résolution simplifiée
        # [n    sum_x  sum_x2][c]   [sum_y  ]
        # [sum_x sum_x2 sum_x3][b] = [sum_xy ]
        # [sum_x2 sum_x3 sum_x4][a]   [sum_x2y]
        
        det = n * (sum_x2 * sum_x4 - sum_x3**2) - sum_x * (sum_x * sum_x4 - sum_x2 * sum_x3) + sum_x2 * (sum_x * sum_x3 - sum_x2**2)
        
        if abs(det) < 1e-10:
            # Matrice singulière, retour à la régression linéaire
            m, b = _linear_regression_coeffs(y_list)
            return [b, m, 0.0]
        
        # Coefficients via règle de Cramer (version simplifiée)
        c = (sum_y * (sum_x2 * sum_x4 - sum_x3**2) - sum_xy * (sum_x * sum_x4 - sum_x2 * sum_x3) + sum_x2y * (sum_x * sum_x3 - sum_x2**2)) / det
        b = (n * (sum_xy * sum_x4 - sum_x2y * sum_x3) - sum_y * (sum_x * sum_x4 - sum_x2 * sum_x3) + sum_x2 * (sum_x * sum_x2y - sum_xy * sum_x2)) / det
        a = (n * (sum_x2 * sum_x2y - sum_xy * sum_x3) - sum_x * (sum_x * sum_x2y - sum_y * sum_x3) + sum_y * (sum_x * sum_x3 - sum_x2**2)) / det
        
        return [c, b, a]
    else:
        # Retour à la régression linéaire
        m, b = _linear_regression_coeffs(y_list)
        return [b, m]


def _moving_average(y: Iterable[float], window: int = 3) -> List[float]:
    """Calcule la moyenne mobile pour lisser les données."""
    y_list = list(float(v) for v in y)
    if len(y_list) < window:
        return y_list
    
    smoothed = []
    for i in range(len(y_list)):
        if i < window - 1:
            smoothed.append(y_list[i])
        else:
            avg = sum(y_list[i-window+1:i+1]) / window
            smoothed.append(avg)
    
    return smoothed


def predict_capacity(data: List[Tuple[str, float]], periods: int = 12, method: str = "linear") -> Optional[List[float]]:
    """Prédit `periods` points futurs selon la méthode choisie.

    Args:
        data: liste de tuples (date, valeur)
        periods: nombre de pas à prédire  
        method: "linear", "polynomial", "moving_average"
    
    Returns:
        Liste des valeurs prédites ou None si impossible
    """
    if not data:
        return None

    # On n'utilise pas la date dans ce modèle simple; on respecte seulement l'ordre
    y_series = [float(v) for _, v in data]
    
    if method == "polynomial":
        coeffs = _polynomial_regression_coeffs(y_series, degree=2)
        n = len(y_series)
        
        if len(coeffs) == 3:  # Quadratique
            c, b, a = coeffs
            preds = [a * (n + i)**2 + b * (n + i) + c for i in range(periods)]
        else:  # Linéaire en fallback
            b, m = coeffs[0], coeffs[1] if len(coeffs) > 1 else 0
            preds = [m * (n + i) + b for i in range(periods)]
            
    elif method == "moving_average":
        # Moyenne mobile avec projection de la tendance
        smoothed = _moving_average(y_series, window=min(5, len(y_series)))
        if len(smoothed) >= 2:
            # Tendance basée sur les derniers points
            trend = (smoothed[-1] - smoothed[-2]) if len(smoothed) >= 2 else 0
            last_value = smoothed[-1]
            preds = [last_value + trend * (i + 1) for i in range(periods)]
        else:
            preds = [smoothed[-1] if smoothed else 0] * periods
            
    else:  # linear (default)
        m, b = _linear_regression_coeffs(y_series)
        n = len(y_series)
        preds = [m * (n + i) + b for i in range(periods)]
    
    # Assurer que les prédictions restent dans des limites raisonnables (0-100% pour l'utilisation)
    preds = [max(0.0, min(100.0, p)) for p in preds]
    
    return preds


def predict_interface_utilization(equipment_id: int, interface_name: str, 
                                days_history: int = 30, periods: int = 12,
                                method: str = "linear") -> Dict[str, Any]:
    """Prédit l'utilisation future d'une interface spécifique.
    
    Args:
        equipment_id: ID de l'équipement
        interface_name: Nom de l'interface
        days_history: Nombre de jours d'historique à utiliser
        periods: Nombre de périodes à prédire
        method: Méthode de prédiction
    
    Returns:
        Dictionnaire avec prédictions et métadonnées
    """
    history = load_interface_history(equipment_id, days_history)
    
    # Filtrer les données pour l'interface spécifique
    interface_data = []
    for entry in history:
        iface_data = entry.get('interface_data', {})
        if interface_name in iface_data:
            timestamp = entry['timestamp']
            utilization = iface_data[interface_name].get('utilization_out', 0)
            interface_data.append((timestamp, utilization))
    
    if not interface_data:
        return {
            'success': False,
            'message': f'Aucune donnée historique pour l\'interface {interface_name}',
            'predictions': []
        }
    
    # Trier par timestamp
    interface_data.sort(key=lambda x: x[0])
    
    # Prédiction
    predictions = predict_capacity(interface_data, periods, method)
    
    if predictions is None:
        return {
            'success': False,
            'message': 'Impossible de générer des prédictions',
            'predictions': []
        }
    
    # Calcul de la tendance
    values = [v for _, v in interface_data]
    trend = "stable"
    if len(values) >= 2:
        recent_avg = sum(values[-5:]) / min(5, len(values))
        older_avg = sum(values[:5]) / min(5, len(values))
        if recent_avg > older_avg * 1.1:
            trend = "increasing"
        elif recent_avg < older_avg * 0.9:
            trend = "decreasing"
    
    # Détection des seuils d'alerte
    critical_periods = []
    for i, pred in enumerate(predictions):
        if pred >= 85.0:
            critical_periods.append(i + 1)
    
    return {
        'success': True,
        'equipment_id': equipment_id,
        'interface_name': interface_name,
        'method': method,
        'historical_points': len(interface_data),
        'trend': trend,
        'current_utilization': values[-1] if values else 0,
        'predictions': predictions,
        'critical_periods': critical_periods,
        'saturation_risk': len(critical_periods) > 0,
        'time_to_saturation': critical_periods[0] if critical_periods else None
    }


def predict_domain_capacity(domain: NetworkDomain, days_history: int = 30,
                          periods: int = 12) -> Dict[str, Any]:
    """Prédit l'évolution de capacité pour tout un domaine réseau."""
    from .inventory.store import get_equipments_by_domain
    
    equipments = get_equipments_by_domain(domain)
    domain_predictions = []
    total_interfaces = 0
    critical_interfaces = 0
    
    for equipment in equipments:
        for interface in equipment.interfaces:
            total_interfaces += 1
            
            prediction = predict_interface_utilization(
                equipment.id, interface.name, days_history, periods
            )
            
            if prediction['success'] and prediction['saturation_risk']:
                critical_interfaces += 1
                domain_predictions.append({
                    'equipment_name': equipment.name,
                    'interface_name': interface.name,
                    'prediction': prediction
                })
    
    # Calcul des métriques agrégées du domaine
    if domain_predictions:
        avg_time_to_saturation = sum(
            p['prediction']['time_to_saturation'] 
            for p in domain_predictions 
            if p['prediction']['time_to_saturation']
        ) / len([p for p in domain_predictions if p['prediction']['time_to_saturation']])
    else:
        avg_time_to_saturation = None
    
    return {
        'domain': domain.value,
        'total_interfaces': total_interfaces,
        'critical_interfaces': critical_interfaces,
        'risk_percentage': (critical_interfaces / total_interfaces * 100) if total_interfaces > 0 else 0,
        'average_time_to_saturation': avg_time_to_saturation,
        'interface_predictions': domain_predictions,
        'recommendations': _generate_capacity_recommendations(domain, critical_interfaces, total_interfaces)
    }


def _generate_capacity_recommendations(domain: NetworkDomain, critical_interfaces: int, 
                                     total_interfaces: int) -> List[Dict[str, str]]:
    """Génère des recommandations basées sur l'analyse prédictive."""
    recommendations = []
    risk_percentage = (critical_interfaces / total_interfaces * 100) if total_interfaces > 0 else 0
    
    if risk_percentage >= 50:
        recommendations.append({
            'priority': 'urgent',
            'action': 'immediate_capacity_upgrade',
            'description': f'Mise à niveau urgente requise - {risk_percentage:.1f}% des interfaces à risque'
        })
    elif risk_percentage >= 25:
        recommendations.append({
            'priority': 'high',
            'action': 'capacity_planning',
            'description': f'Planification de capacité nécessaire - {risk_percentage:.1f}% des interfaces à risque'
        })
    elif risk_percentage >= 10:
        recommendations.append({
            'priority': 'medium',
            'action': 'monitoring_increase',
            'description': f'Surveillance renforcée recommandée - {risk_percentage:.1f}% des interfaces à risque'
        })
    
    # Recommandations spécifiques par domaine
    if domain == NetworkDomain.CORE_INTERNET and critical_interfaces > 0:
        recommendations.append({
            'priority': 'critical',
            'action': 'redundancy_check',
            'description': 'Vérifier la redondance des liens critiques'
        })
    elif domain == NetworkDomain.BACKBONE and critical_interfaces > 2:
        recommendations.append({
            'priority': 'high',
            'action': 'load_balancing',
            'description': 'Optimisation de la répartition de charge'
        })
    
    return recommendations


def generate_capacity_report(periods: int = 12, format: str = "json") -> Dict[str, Any]:
    """Génère un rapport complet d'analyse prédictive."""
    report = {
        'generated_at': datetime.now().isoformat(),
        'prediction_periods': periods,
        'domains': {}
    }
    
    # Analyse par domaine
    for domain in NetworkDomain:
        domain_analysis = predict_domain_capacity(domain, periods=periods)
        report['domains'][domain.value] = domain_analysis
    
    # Résumé global
    total_interfaces = sum(d['total_interfaces'] for d in report['domains'].values())
    total_critical = sum(d['critical_interfaces'] for d in report['domains'].values())
    
    report['summary'] = {
        'total_interfaces': total_interfaces,
        'critical_interfaces': total_critical,
        'global_risk_percentage': (total_critical / total_interfaces * 100) if total_interfaces > 0 else 0,
        'highest_risk_domain': max(
            report['domains'].items(), 
            key=lambda x: x[1]['risk_percentage']
        )[0] if report['domains'] else None
    }
    
    return report


def export_predictions_excel(predictions: Dict[str, Any], filename: str = None) -> str:
    """Exporte les prédictions vers un fichier Excel."""
    if filename is None:
        filename = f"predictions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    try:
        import openpyxl
        from openpyxl.workbook import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Analyse Prédictive"
        
        # En-têtes
        headers = ['Domaine', 'Équipement', 'Interface', 'Utilisation Actuelle', 
                   'Tendance', 'Risque Saturation', 'Périodes à Risque']
        ws.append(headers)
        
        # Mise en forme des en-têtes
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Données
        for domain_name, domain_data in predictions.get('domains', {}).items():
            for interface_pred in domain_data.get('interface_predictions', []):
                pred = interface_pred['prediction']
                ws.append([
                    domain_name,
                    interface_pred['equipment_name'],
                    interface_pred['interface_name'],
                    f"{pred['current_utilization']:.1f}%",
                    pred['trend'],
                    "Oui" if pred['saturation_risk'] else "Non",
                    str(pred['critical_periods']) if pred['critical_periods'] else "Aucun"
                ])
        
        wb.save(filename)
        return filename
        
    except ImportError:
        # Fallback vers CSV si openpyxl n'est pas disponible
        import csv
        csv_filename = filename.replace('.xlsx', '.csv')
        
        with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Domaine', 'Équipement', 'Interface', 'Utilisation Actuelle', 
                           'Tendance', 'Risque Saturation', 'Périodes à Risque'])
            
            for domain_name, domain_data in predictions.get('domains', {}).items():
                for interface_pred in domain_data.get('interface_predictions', []):
                    pred = interface_pred['prediction']
                    writer.writerow([
                        domain_name,
                        interface_pred['equipment_name'],
                        interface_pred['interface_name'],
                        f"{pred['current_utilization']:.1f}%",
                        pred['trend'],
                        "Oui" if pred['saturation_risk'] else "Non",
                        str(pred['critical_periods']) if pred['critical_periods'] else "Aucun"
                    ])
        
        return csv_filename
