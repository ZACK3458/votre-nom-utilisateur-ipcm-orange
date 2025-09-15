
from flask import render_template, redirect, url_for, send_from_directory, jsonify, request, Response
import time
from app import app
from app.inventory.store import (
    load_inventory, add_equipment, update_equipment, delete_equipment,
    load_equipment_objects, get_equipment_by_id, save_interface_history
)
from app.inventory.domains import domain_manager
from app.inventory.models import NetworkDomain, EquipmentType
from app.snmp.collector import SNMPCollector, collect_interface_data
from app.predictive import (
    predict_interface_utilization, predict_domain_capacity, 
    generate_capacity_report, export_predictions_excel
)
import csv
from io import StringIO
from datetime import datetime, timedelta
try:
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception:  # keep offline even if openpyxl missing
    openpyxl = None

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(app.static_folder, 'favicon.ico', mimetype='image/vnd.microsoft.icon')

@app.route('/users')
def users():
    return render_template('user_space/user_space.html')

@app.route('/inventory')
def inventory():
    items = load_inventory()
    # Enrichir avec les données du domaine manager
    domain_summary = domain_manager.get_domain_summary()
    return render_template('inventory.html', items=items, domain_summary=domain_summary)

@app.route('/inventory/add', methods=['POST'])
def inventory_add():
    data = request.get_json(silent=True) or request.form.to_dict()
    eq = add_equipment(data)
    return jsonify(eq), 201

@app.route('/inventory/<int:equip_id>', methods=['PATCH'])
def inventory_update(equip_id: int):
    changes = request.get_json(silent=True) or request.form.to_dict()
    ok = update_equipment(equip_id, changes)
    return jsonify({'updated': ok}), (200 if ok else 404)

@app.route('/inventory/<int:equip_id>', methods=['DELETE'])
def inventory_delete(equip_id: int):
    ok = delete_equipment(equip_id)
    return jsonify({'deleted': ok}), (200 if ok else 404)

@app.route('/inventory/export.csv')
def inventory_export_csv():
    items = load_inventory()
    si = StringIO()
    writer = csv.DictWriter(si, fieldnames=[
        'id','name','type','brand','model','software_version','ip_address',
        'location','domain','support_status','modules','criticality_level',
        'installation_date','serial_number','asset_tag'
    ])
    writer.writeheader()
    for it in items:
        # Adapter les nouvelles colonnes
        row_data = {k: it.get(k, '') for k in writer.fieldnames}
        if isinstance(row_data.get('modules'), list):
            row_data['modules'] = ', '.join(row_data['modules'])
        writer.writerow(row_data)
    output = si.getvalue()
    return Response(output, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename="inventory.csv"'})

@app.route('/inventory/export.xlsx')
def inventory_export_xlsx():
    if openpyxl is None:
        return jsonify({'error': 'export xlsx indisponible (openpyxl manquant)'}), 503
    
    items = load_inventory()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventaire IPCM"
    
    # En-têtes avec style
    headers = ['ID','Nom','Type','Marque','Modèle','Version SW','IP','Localisation',
               'Domaine','Statut Support','Modules','Criticité','Date Installation',
               'Numéro Série','Asset Tag','Contrat Support']
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Données
    for it in items:
        modules_str = ', '.join(it.get('modules', [])) if isinstance(it.get('modules'), list) else it.get('modules', '')
        ws.append([
            it.get('id', ''),
            it.get('name', ''),
            it.get('type', ''),
            it.get('brand', ''),
            it.get('model', ''),
            it.get('software_version', ''),
            it.get('ip_address', ''),
            it.get('location', ''),
            it.get('domain', ''),
            it.get('support_status', ''),
            modules_str,
            it.get('criticality_level', ''),
            it.get('installation_date', ''),
            it.get('serial_number', ''),
            it.get('asset_tag', ''),
            it.get('vendor_support_contract', '')
        ])
    
    # Auto-ajustement des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(bio.read(), 
                   mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                   headers={'Content-Disposition': 'attachment; filename="inventaire_ipcm.xlsx"'})

# Routes API pour la collecte SNMP
@app.route('/api/snmp/collect/<int:equipment_id>')
def snmp_collect_equipment(equipment_id: int):
    """Collecte SNMP pour un équipement spécifique."""
    equipment = get_equipment_by_id(equipment_id)
    if not equipment:
        return jsonify({'error': 'Équipement non trouvé'}), 404
    
    if not equipment.snmp_enabled:
        return jsonify({'error': 'SNMP désactivé pour cet équipement'}), 400
    
    collector = SNMPCollector(equipment.ip_address, equipment.snmp_community)
    
    if not collector.test_connectivity():
        return jsonify({'error': 'Impossible de se connecter via SNMP'}), 500
    
    interfaces = collector.collect_all_interfaces()
    
    # Mettre à jour l'équipement avec les nouvelles données d'interfaces
    equipment.interfaces = interfaces
    equipments = load_equipment_objects()
    for i, eq in enumerate(equipments):
        if eq.id == equipment_id:
            equipments[i] = equipment
            break
    
    from app.inventory.store import save_equipment_objects
    save_equipment_objects(equipments)
    
    # Sauvegarder l'historique pour l'analyse prédictive
    interface_data = {
        iface.name: {
            'utilization_in': iface.in_utilization,
            'utilization_out': iface.out_utilization,
            'status': iface.status.value,
            'speed': iface.speed
        }
        for iface in interfaces
    }
    save_interface_history(equipment_id, interface_data)
    
    return jsonify({
        'success': True,
        'equipment': equipment.name,
        'interfaces_collected': len(interfaces),
        'interfaces': [
            {
                'name': iface.name,
                'status': iface.status.value,
                'utilization_in': iface.in_utilization,
                'utilization_out': iface.out_utilization,
                'speed': iface.speed
            }
            for iface in interfaces
        ]
    })

@app.route('/api/domains/<domain>/metrics')
def domain_metrics(domain: str):
    """API pour récupérer les métriques d'un domaine."""
    try:
        domain_enum = NetworkDomain(domain)
        metrics = domain_manager.get_domain_metrics(domain_enum)
        alerts = domain_manager.check_domain_alerts(domain_enum)
        
        return jsonify({
            'domain': domain,
            'metrics': {
                'total_equipments': metrics.total_equipments,
                'active_equipments': metrics.active_equipments,
                'critical_equipments': metrics.critical_equipments,
                'total_interfaces': metrics.total_interfaces,
                'active_interfaces': metrics.active_interfaces,
                'average_utilization_in': metrics.average_utilization_in,
                'average_utilization_out': metrics.average_utilization_out,
                'peak_utilization': metrics.peak_utilization,
                'equipments_by_type': metrics.equipments_by_type,
                'support_status_distribution': metrics.support_status_distribution
            },
            'alerts': alerts
        })
    except ValueError:
        return jsonify({'error': 'Domaine non valide'}), 400

@app.route('/api/predictive/interface/<int:equipment_id>/<interface_name>')
def predict_interface(equipment_id: int, interface_name: str):
    """API pour la prédiction d'utilisation d'une interface."""
    days = request.args.get('days', 30, type=int)
    periods = request.args.get('periods', 12, type=int)
    method = request.args.get('method', 'linear')
    
    prediction = predict_interface_utilization(
        equipment_id, interface_name, days, periods, method
    )
    
    return jsonify(prediction)

@app.route('/api/predictive/domain/<domain>')
def predict_domain(domain: str):
    """API pour la prédiction de capacité d'un domaine."""
    try:
        domain_enum = NetworkDomain(domain)
        days = request.args.get('days', 30, type=int)
        periods = request.args.get('periods', 12, type=int)
        
        prediction = predict_domain_capacity(domain_enum, days, periods)
        return jsonify(prediction)
    except ValueError:
        return jsonify({'error': 'Domaine non valide'}), 400

@app.route('/api/predictive/report')
def predictive_report():
    """Génère un rapport complet d'analyse prédictive."""
    periods = request.args.get('periods', 12, type=int)
    format_type = request.args.get('format', 'json')
    
    report = generate_capacity_report(periods, format_type)
    
    if format_type == 'excel':
        filename = export_predictions_excel(report)
        return jsonify({
            'success': True,
            'message': 'Rapport Excel généré',
            'filename': filename
        })
    
    return jsonify(report)

@app.route('/snmp')
def snmp():
    return render_template('interfaces/interfaces.html')

@app.route('/dashboard')
def dashboard():
    # Données enrichies avec les nouvelles fonctionnalités
    domain_summary = domain_manager.get_domain_summary()
    equipments = load_equipment_objects()[:5]  # Top 5 pour le dashboard
    
    # KPIs principaux
    kpis = {
        'total_equipments': domain_summary['total_equipments'],
        'total_interfaces': domain_summary['total_interfaces'],
        'active_interfaces': domain_summary['active_interfaces'],
        'critical_alerts': domain_summary['critical_alerts'],
        'warning_alerts': domain_summary['warning_alerts']
    }
    
    # Interfaces critiques toutes domaines confondues
    critical_interfaces = []
    for domain in NetworkDomain:
        critical_interfaces.extend(
            domain_manager.get_critical_interfaces_by_domain(domain)[:3]
        )
    
    # Alertes récentes
    recent_alerts = domain_summary['alerts'][:10]
    
    return render_template('dashboard.html', 
                         equipments=[eq.to_dict() for eq in equipments],
                         domain_summary=domain_summary,
                         kpis=kpis,
                         critical_interfaces=critical_interfaces,
                         recent_alerts=recent_alerts)

@app.route('/reporting')
def reporting():
    domain_summary = domain_manager.get_domain_summary()
    return render_template('reporting.html', domain_summary=domain_summary)

@app.route('/predictive')
def predictive():
    # Données pour l'interface prédictive
    domains_data = {}
    for domain in NetworkDomain:
        prediction = predict_domain_capacity(domain)
        domains_data[domain.value] = prediction
    
    return render_template('predictive.html', domains_data=domains_data)

@app.route('/security')
def security():
    return render_template('security.html')

@app.route('/features')
def features():
    return render_template('features.html')

@app.route('/roadmap')
def roadmap():
    return render_template('roadmap/roadmap.html')

@app.route('/interfaces')
def interfaces():
    # Page interfaces enrichie avec données SNMP
    equipments = load_equipment_objects()
    interfaces_data = []
    
    for eq in equipments:
        for iface in eq.interfaces:
            interfaces_data.append({
                'equipment_name': eq.name,
                'equipment_ip': eq.ip_address,
                'domain': eq.domain.value,
                'interface_name': iface.name,
                'interface_description': iface.description,
                'status': iface.status.value,
                'speed': iface.speed,
                'utilization_in': iface.in_utilization,
                'utilization_out': iface.out_utilization,
                'last_updated': iface.last_updated.isoformat() if iface.last_updated else None
            })
    
    return render_template('interfaces/interfaces.html', interfaces=interfaces_data)

@app.route('/user-space')
def user_space():
    return render_template('user_space/user_space.html')

@app.route('/service')
def service():
    return render_template('service/service.html')

@app.route('/admin-data')
def admin_data():
    return render_template('admin-data.html')

@app.route('/plan-adressage')
def plan_adressage():
    return render_template('plan-adressage.html')

@app.route('/architecture')
def architecture():
    return render_template('architecture.html')

@app.route('/healthz')
def healthz():
    """Endpoint de santé simple pour les probes/monitoring.
    Retourne un JSON minimal indiquant le bon fonctionnement.
    """
    return jsonify(status='ok'), 200

@app.route('/metrics')
def metrics():
    """Expose des métriques basiques pour supervision/light observability."""
    uptime = time.time() - getattr(app, 'start_time', time.time())
    routes_count = len(app.url_map._rules)
    
    # Métriques IPCM enrichies
    domain_summary = domain_manager.get_domain_summary()
    
    return jsonify({
        'service': app.config.get('SERVICE_NAME', 'ipcm'),
        'version': app.config.get('VERSION', '0.0.0'),
        'uptime_s': round(uptime, 3),
        'routes_count': routes_count,
        'status': 'ok',
        'ipcm_metrics': {
            'total_equipments': domain_summary['total_equipments'],
            'total_interfaces': domain_summary['total_interfaces'],
            'active_interfaces': domain_summary['active_interfaces'],
            'critical_alerts': domain_summary['critical_alerts'],
            'warning_alerts': domain_summary['warning_alerts']
        }
    }), 200

# Extra routes referenced by navbar
@app.route('/precablage')
def precablage():
    # Placeholder page mapped to existing service template
    return render_template('service/service.html')

@app.route('/journal')
def journal():
    # Placeholder: reuse reporting layout for activity journal
    return render_template('reporting.html')

@app.route('/logout')
def logout():
    # Offline mode: just redirect to dashboard
    return redirect(url_for('index'))

@app.route('/')
def index():
    # Page d'accueil enrichie avec tableau de bord complet
    return redirect(url_for('dashboard'))
